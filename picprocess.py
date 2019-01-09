import os
import openpyxl
import shutil


def picprocess():
    root_dir=r'D:\picprocess'
    result_dir=os.path.join(root_dir,'结果')
    log_path=os.path.join(root_dir,'处理结果.txt')
    goods_dict={}
    if not os.path.isdir(result_dir):
        os.makedirs(result_dir)
        print('初始化完成....')
        print('\n')
        print('请将excel文件以及图片文件放到目录 {}下，再重新执行此脚本'.format(root_dir))
        input()
        return
    if os.listdir(root_dir)==['结果']:
        print('请将excel文件以及图片文件放到目录 {}下，再重新执行此脚本'.format(root_dir))
        input()
        return
    
    for root_file_name in os.listdir(root_dir):
        filetype=root_file_name.split('.')[-1]
        file_path=os.path.join(root_dir,root_file_name)
        if filetype=='xlsx' or filetype=='xls':
            try:
                workbook=openpyxl.load_workbook(file_path)
            except PermissionError:
                print('请关闭打开的 {} 文件，再重新执行脚本，谢谢'.format(root_file_name))
                input()
            wst=workbook.active
            rows=wst.max_row
            for i in range(2,rows+1):
                goods_dict[wst.cell(row=i,column=1).value]=wst.cell(row=i,column=2).value


    for value in goods_dict.values():
        kuanse_dir=os.path.join(result_dir,value)
        if not os.path.isdir(kuanse_dir):
            os.makedirs(kuanse_dir)
    
    tmp=[]

    i=0
    for file_name in os.listdir(root_dir):
        if file_name!='结果' and os.path.isdir(os.path.join(root_dir,file_name)):
            pic_dir=os.path.join(root_dir,file_name)
            for pic_name in [name for name in os.listdir(pic_dir) if name.split('.')[-1]=='jpg' or name.split('.')[-1]=='png']:
                goodscode=pic_name.split('.')[0]
                try:
                    kuanse=goods_dict[goodscode]
                    old_path=os.path.join(pic_dir,pic_name)
                    new_path=os.path.join(result_dir,kuanse,pic_name)
                    shutil.copyfile(old_path,new_path)
                    i+=1
                except KeyError:
                    tmp.append(goodscode)

    for file_dir in os.listdir(result_dir):
        if not os.listdir(os.path.join(result_dir,file_dir)):
            os.removedirs(os.path.join(result_dir,file_dir))

    s='处理成功，共处理好了{}个款色，部分命名图片的货号在excel里没有找到，这些货号见 {}'.format(i,log_path)
    print(s)
    with open(log_path,'w',encoding='utf-8') as f:
        f.write('处理成功，共处理好了{}个款色，部分命名图片的货号在excel里没有找到,这些货号如下： \n{}'.format(i,'\n'.join(tmp)))
    input()



picprocess()