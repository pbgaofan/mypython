#-- coding:utf-8 --
#!/usr/bin/python
'''''
Created on 2017-11-01
@author: honest

'''
import os
import io
import sys
import xlwt
import time
from xlutils.copy import copy
from urllib import request,parse
from bs4 import BeautifulSoup
from datetime import datetime
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='utf-8')

def search():
    base_url='http://www.ip138.com:8080/search.asp?'
    excel_dir=r'D:\myexe'
    temp_excel_path=os.path.join(excel_dir,'输入.xlsx')
    if os.path.isdir(excel_dir):
        if os.path.isfile(temp_excel_path)==False:
            workbook=openpyxl.Workbook()
            ws=workbook.active
            lis=['手机号','15824227833','13811652274']
            for i in range(1,4):
                ws.cell(row=i,column=1).value=lis[i-1]
            ws.cell(row=1,column=2).value='归属地'
            workbook.save(temp_excel_path)
        else:
            wbk=openpyxl.load_workbook(temp_excel_path)
            ws=wbk.active
            length=ws.max_row
            k=2
            while length>10000:
                for i in range(k,k+10000):
                    phone=ws.cell(row=i,column=1).value
                    dic = {
                        'mobile': phone,
                        'action': 'mobile'
                    }
                    url = base_url + parse.urlencode(dic)
                    r = request.urlopen(url)
                    soup = BeautifulSoup(r.read().decode('gb2312'), 'html.parser')
                    content = soup.select('td')[6].text
                    ws.cell(row=i,column=2).value=content
                wbk.save(temp_excel_path)
                length-=10000
                k+=10000
                wbk=openpyxl.load_workbook(temp_excel_path)
                ws=wbk.active
            for i in range(k,k+length-1):
                phone=ws.cell(row=i,column=1).value
                dic = {
                    'mobile': phone,
                    'action': 'mobile'
                }
                url = base_url + parse.urlencode(dic)
                r = request.urlopen(url)
                soup = BeautifulSoup(r.read().decode('gb2312'), 'html.parser')
                content = soup.select('td')[6].text
                ws.cell(row=i,column=2).value=content
            wbk.save(temp_excel_path)
            result_path=os.path.join(excel_dir,'输出结果.xlsx')
            os.rename(temp_excel_path,result_path)
    else:
        os.makedirs(excel_dir)
        search()
search()
